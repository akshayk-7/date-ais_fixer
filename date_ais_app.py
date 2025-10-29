import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from openpyxl import load_workbook
import customtkinter as ctk

# ---------- Utility ----------
def find_header_row(path, max_scan=30, sheet_name=0):
    raw = pd.read_excel(path, sheet_name=sheet_name, header=None, dtype=str)
    max_scan = min(max_scan, len(raw))
    want = {"stock name", "buy date", "buydate", "sell date", "selldate", "quantity"}
    for i in range(max_scan):
        row = raw.iloc[i].astype(str).str.strip().str.lower().tolist()
        if any(w in row for w in want):
            return i
    return 0

def convert_date_series(series: pd.Series) -> pd.Series:
    s = series.astype(str).str.strip()
    s8 = s.str.extract(r'(\d{8})', expand=False)
    ymd = pd.to_datetime(s8, format="%Y%m%d", errors="coerce")
    dmy = pd.to_datetime(s8, format="%d%m%Y", errors="coerce")
    parsed = ymd.fillna(dmy)
    fallback = pd.to_datetime(s, dayfirst=True, errors="coerce")
    parsed = parsed.fillna(fallback)
    if parsed.notna().sum() >= max(1, int(len(series) * 0.3)):
        return parsed
    else:
        return series

# ---------- Per-file tab container ----------
class FileTab:
    """Holds state and UI for a single file inside its own tab."""
    def __init__(self, parent_notebook, file_path, app_ref):
        self.app = app_ref
        self.file_path = file_path
        self.file_name = file_path.split("/")[-1].split("\\")[-1]

        # data
        self.dfs = {}
        self.date_cols = {}
        self.current_sheet = None
        self.preview_df = None

        # UI: create tab frame
        self.tab_frame = tk.Frame(parent_notebook)
        parent_notebook.add(self.tab_frame, text=self.file_name)

        # --- Top controls for this tab ---
        top = tk.Frame(self.tab_frame, pady=6)
        top.pack(fill="x")

        tk.Label(top, text="Sheet:", font=("Arial", 10, "bold")).pack(side="left", padx=(4, 4))
        self.sheet_selector = ttk.Combobox(top, state="readonly", width=28)
        self.sheet_selector.bind("<<ComboboxSelected>>", self.switch_sheet)
        self.sheet_selector.pack(side="left", padx=(0, 10))

        tk.Label(top, text="🔍 Search:", font=("Arial", 10)).pack(side="left", padx=(10, 4))
        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", self.apply_search)
        self.search_entry = ctk.CTkEntry(top, textvariable=self.search_var,
                                         placeholder_text="Type to search...",
                                         width=220, height=32, corner_radius=12)
        self.search_entry.pack(side="left", padx=(0, 10))

        self.export_btn = tk.Button(top, text="💾 Export This File",
                                    command=self.export_this_file, font=("Arial", 11),
                                    width=18, bg="#58D68D")
        self.export_btn.pack(side="left", padx=6)

        # NEW: Add Long/Short Flag Button
        self.flag_btn = tk.Button(top, text="➕ Add Long/Short Flags",
                                  command=self.add_flags,
                                  font=("Arial", 11), width=22, bg="#F7DC6F")
        self.flag_btn.pack(side="left", padx=6)

        # --- Table area ---
        table_wrap = tk.Frame(self.tab_frame)
        table_wrap.pack(expand=True, fill="both", padx=8, pady=6)

        self.tree = ttk.Treeview(table_wrap, show="headings")
        vsb = ttk.Scrollbar(table_wrap, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(table_wrap, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscroll=vsb.set, xscroll=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        table_wrap.grid_rowconfigure(0, weight=1)
        table_wrap.grid_columnconfigure(0, weight=1)

        # copy/paste bindings
        self.tree.bind("<Control-c>", self.copy_selection)
        self.tree.bind("<Control-C>", self.copy_selection)
        self.tree.bind("<Control-v>", self.paste_selection)
        self.tree.bind("<Control-V>", self.paste_selection)

        # right-click header menu
        self.tree.bind("<Button-3>", self.show_header_menu)
        self.header_menu = tk.Menu(self.tree, tearoff=0)
        self.header_menu.add_command(label="📋 Copy Entire Column", command=self.copy_entire_column)
        self.header_clicked_col = None

        # --- Status (per tab) ---
        self.status = tk.Label(self.tab_frame, text="Load a file...", font=("Arial", 10), fg="gray")
        self.status.pack(pady=(0, 6))

    # ---- Data load ----
    def load(self):
        xl = pd.ExcelFile(self.file_path)
        total = len(xl.sheet_names)
        for idx, sheet in enumerate(xl.sheet_names, start=1):
            self.app.show_progress(f"Loading {self.file_name} → {sheet}  {idx}/{total}", idx, total)
            hdr = find_header_row(self.file_path, sheet_name=sheet)
            df = pd.read_excel(self.file_path, sheet_name=sheet, header=hdr)

            try:
                if (df.iloc[0].astype(str).str.lower().values ==
                    df.columns.astype(str).str.lower().values).all():
                    df = df.drop(df.index[0]).reset_index(drop=True)
            except Exception:
                pass

            self.date_cols[sheet] = []
            for col in df.columns:
                before = df[col]
                after = convert_date_series(before)
                df[col] = after
                if pd.api.types.is_datetime64_any_dtype(df[col]):
                    self.date_cols[sheet].append(col)

            self.dfs[sheet] = df

        self.sheet_selector["values"] = list(self.dfs.keys())
        if self.dfs:
            self.sheet_selector.current(0)
            self.switch_sheet()
            self.status.config(text=f"✔ Loaded {self.file_name} with {len(self.dfs)} sheet(s).", fg="green")

    # ---- Add Long/Short term flags ----
    def add_flags(self):
        if not self.current_sheet or self.current_sheet not in self.dfs:
            messagebox.showerror("Error", "No sheet selected!")
            return
        df = self.dfs[self.current_sheet]

        if "Asset Type" not in df.columns:
            messagebox.showerror("Error", "'Asset Type' column not found in this sheet!")
            return

        # Insert "Flag" right after Asset Type
        asset_idx = df.columns.get_loc("Asset Type") + 1
        if "Flag" not in df.columns:
            df.insert(asset_idx, "Flag",
                      df["Asset Type"].apply(lambda x: "yes" if str(x).strip().lower() == "long term"
                                             else "no" if str(x).strip().lower() == "short term"
                                             else ""))
        self.dfs[self.current_sheet] = df
        self.switch_sheet()
        self.status.config(text="✅ Flags added next to Asset Type.", fg="green")

    # ---- Right-click menu handlers ----
    def show_header_menu(self, event):
        region = self.tree.identify("region", event.x, event.y)
        if region == "heading":
            col_id = self.tree.identify_column(event.x)
            col_index = int(col_id.replace("#", "")) - 1
            if col_index >= 0 and col_index < len(self.tree["columns"]):
                self.header_clicked_col = self.tree["columns"][col_index]
                self.header_menu.post(event.x_root, event.y_root)

    def copy_entire_column(self):
        if not self.header_clicked_col:
            return
        col_index = self.tree["columns"].index(self.header_clicked_col)
        values = []
        for row_id in self.tree.get_children():
            row = self.tree.item(row_id, "values")
            if col_index < len(row):
                values.append(str(row[col_index]))
        text = "\n".join(values)
        self.tree.clipboard_clear()
        self.tree.clipboard_append(text)
        self.tree.update()
        messagebox.showinfo("Copied", f"Copied {len(values)} rows from '{self.header_clicked_col}' column!")

    # ---- UI helpers ----
    def switch_sheet(self, event=None):
        sheet = self.sheet_selector.get()
        if not sheet:
            return
        self.current_sheet = sheet
        df = self.dfs[sheet].copy()
        for col in self.date_cols.get(sheet, []):
            df[col] = pd.to_datetime(df[col], errors="coerce").dt.strftime("%d-%m-%Y")
        self.show_preview(df)

    def show_preview(self, df):
        self.preview_df = df
        self.tree.delete(*self.tree.get_children())
        self.tree["columns"] = list(df.columns)
        for col in df.columns:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=140, anchor="center")
        for i, (_, row) in enumerate(df.head(500).iterrows()):
            tag = "evenrow" if i % 2 == 0 else "oddrow"
            self.tree.insert("", "end",
                             values=[("" if pd.isna(v) else v) for v in row.tolist()],
                             tags=(tag,))
        self.tree.tag_configure("evenrow", background="#F8F9F9")
        self.tree.tag_configure("oddrow", background="#EBF5FB")

    def apply_search(self, *args):
        if self.current_sheet is None or self.preview_df is None:
            return
        keyword = self.search_var.get().lower().strip()
        if keyword == "":
            df_filtered = self.preview_df
        else:
            mask = self.preview_df.apply(
                lambda row: row.astype(str).str.lower().str.contains(keyword).any(),
                axis=1
            )
            df_filtered = self.preview_df[mask]
        self.tree.delete(*self.tree.get_children())
        for i, (_, row) in enumerate(df_filtered.head(500).iterrows()):
            tag = "evenrow" if i % 2 == 0 else "oddrow"
            self.tree.insert("", "end",
                             values=[("" if pd.isna(v) else v) for v in row.tolist()],
                             tags=(tag,))

    # ---- Copy/Paste ----
    def copy_selection(self, event=None):
        selected = self.tree.selection()
        if not selected:
            return
        rows = []
        for item in selected:
            values = self.tree.item(item, "values")
            rows.append("\t".join(str(v) for v in values))
        text = "\n".join(rows)
        self.tree.clipboard_clear()
        self.tree.clipboard_append(text)
        self.tree.update()

    def paste_selection(self, event=None):
        if not self.current_sheet or self.preview_df is None:
            return
        try:
            text = self.tree.clipboard_get()
            lines = text.splitlines()
            selected = self.tree.selection()
            if not selected:
                return
            for i, item in enumerate(selected):
                if i >= len(lines):
                    break
                values = lines[i].split("\t")
                col_count = len(self.tree["columns"])
                if len(values) < col_count:
                    values += [""] * (col_count - len(values))
                self.tree.item(item, values=values)
                row_index = self.tree.index(item)
                for j, col in enumerate(self.tree["columns"]):
                    self.dfs[self.current_sheet].iloc[row_index, j] = values[j]
            self.status.config(text="✅ Pasted into selected rows.", fg="green")
        except Exception as e:
            messagebox.showerror("Paste Error", str(e))

    # ---- Export ----
    def export_this_file(self):
        if not self.dfs:
            return
        save_path = filedialog.asksaveasfilename(
            title=f"Export {self.file_name} As",
            defaultextension=".xlsx",
            filetypes=[("Excel Files", "*.xlsx")]
        )
        if not save_path:
            return
        try:
            writer = pd.ExcelWriter(save_path, engine="openpyxl")
            total = len(self.dfs)
            for idx, (sheet, df) in enumerate(self.dfs.items(), start=1):
                self.app.show_progress(f"Exporting {self.file_name} → {sheet}  {idx}/{total}", idx, total)
                df.to_excel(writer, sheet_name=sheet, index=False)
            writer.close()
            wb = load_workbook(save_path)
            for sheet, df in self.dfs.items():
                ws = wb[sheet]
                for col_name in self.date_cols.get(sheet, []):
                    col_idx = list(df.columns).index(col_name) + 1
                    for row in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row, column=col_idx)
                        if cell.value is not None:
                            cell.number_format = "DD-MM-YYYY"
            wb.save(save_path)
            self.app.hide_progress()
            messagebox.showinfo("Exported", f"File saved successfully:\n{save_path}")
            self.status.config(text="✅ Export successful!", fg="green")
        except Exception as e:
            self.app.hide_progress()
            messagebox.showerror("Export Error", str(e))
            self.status.config(text="❌ Export failed.", fg="red")

# ---------- Main App ----------
class ExcelDateFixerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("📊 Excel Date Fixer and AIS Editor")
        self.file_tabs = []

        title = tk.Label(root, text="Excel Date Fixer and AIS Editor",
                         font=("Arial", 18, "bold"), bg="#2E86C1", fg="white", pady=10)
        title.pack(fill="x")

        top = tk.Frame(root, pady=8)
        top.pack(fill="x")

        self.load_button = ctk.CTkButton(top, text="📂 Select Excel File(s)",
                                         command=self.load_files,
                                         corner_radius=20, fg_color="#5DADE2", hover_color="#3498DB")
        self.load_button.pack(side="left", padx=10)

        self.files_notebook = ttk.Notebook(root)
        self.files_notebook.pack(expand=True, fill="both", padx=8, pady=6)

        self.progress_frame = tk.Frame(root)
        self.progress = ttk.Progressbar(self.progress_frame, orient="horizontal",
                                        mode="determinate", length=450)
        self.progress.pack(pady=3)
        self.progress_label = tk.Label(self.progress_frame, text="", font=("Arial", 10))
        self.progress_label.pack()

        self.status = tk.Label(root, text="Select one or more Excel files.",
                               font=("Arial", 11), fg="gray")
        self.status.pack(pady=6)

        style = ttk.Style()
        style.configure("Treeview.Heading", font=("Arial", 11, "bold"))
        style.configure("Treeview", font=("Arial", 10), rowheight=22)
        style.map("Treeview", background=[("selected", "#AED6F1")])

    def show_progress(self, text, value, maximum):
        self.progress_frame.pack(pady=5)
        self.progress["maximum"] = maximum
        self.progress["value"] = value
        percent = int((value / maximum) * 100) if maximum else 0
        self.progress_label.config(text=f"{text} ({percent}%)")
        self.root.update_idletasks()

    def hide_progress(self):
        self.progress_frame.pack_forget()

    def load_files(self):
        file_paths = filedialog.askopenfilenames(
            title="Select Excel File(s)",
            filetypes=[("Excel Files", "*.xlsx *.xls")]
        )
        if not file_paths:
            return
        created = 0
        for path in file_paths:
            try:
                tab = FileTab(self.files_notebook, path, self)
                tab.load()
                self.file_tabs.append(tab)
                created += 1
            except Exception as e:
                self.hide_progress()
                messagebox.showerror("Load Error", f"{path}\n\n{e}")
        self.hide_progress()
        if created:
            self.status.config(text=f"✔ Loaded {created} file(s) into tabs.", fg="green")
        else:
            self.status.config(text="No files loaded.", fg="red")
        # --- Footer ---
        footer = tk.Label(root, text="Developed by ANK",
                          font=("Arial", 10, "italic"), fg="gray")
        footer.pack(side="bottom", pady=5)

if __name__ == "__main__":
    root = tk.Tk()
    root.geometry("1200x800")
    app = ExcelDateFixerApp(root)
    root.mainloop()
